import concurrent.futures
import hashlib
import re
import sys
import time
from urllib.parse import urlencode
import requests
from colorama import Fore, init
from openpyxl import Workbook
from dotenv import load_dotenv
import os

load_dotenv()


wb1 = Workbook()
ws1 = wb1["Sheet"]
wb2 = Workbook()
ws2 = wb2["Sheet"]
wb3 = Workbook()
ws3 = wb3["Sheet"]
wb4 = Workbook()
ws4 = wb4["Sheet"]
wb5 = Workbook()
ws5 = wb5["Sheet"]

ws1.append(
    [
        "Project Name",
        "Project Total Unit",
        "Project Available Unit",
        "Unit Name",
        "Name Total Unit",
        "Unit Name Available Unit",
        "Unit Type",
        "Type Total Unit",
        "Unit Type Available Unit",
        "Remark",
    ]
)
ws2.append(
    [
        "Project Name",
        "Project Total Unit",
        "Project Available Unit",
        "Unit Name",
        "Name Total Unit",
        "Unit Name Available Unit",
        "Unit Type",
        "Type Total Unit",
        "Unit Type Available Unit",
        "Remark",
    ]
)
ws3.append(
    [
        "Project Name",
        "Project Total Unit",
        "Project Available Unit",
        "Unit Name",
        "Name Total Unit",
        "Unit Name Available Unit",
        "Unit Type",
        "Type Total Unit",
        "Unit Type Available Unit",
        "Remark",
    ]
)
ws4.append(
    [
        "Project Name",
        "Unit Name",
        "Unit Type",
        "Min Price",
        "Max Price",
        "Min PSF",
        "Max PSF",
        "Remark",
    ]
)
ws5.append(
    [
        "Project Name",
        "Unit Name",
        "Unit Type",
        "Unit Number",
        "Price",
        "PSF",
        "Size",
        "Remark",
    ]
)

init(convert=True)

sess = requests.Session()

sess.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.5060.134 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Content-Type": "application/x-www-form-urlencoded",
    "Origin": "https://app.singmap.com",
    "Referer": "https://app.singmap.com/",
    "Sec-Fetch-Dest": "empty",
    "Sec-Fetch-Mode": "cors",
    "Sec-Fetch-Site": "same-site",
})

params = {
    "agentId": "",
    "brokeId": "",
    "token": "",
    "source": "",
    "timestamp": int(time.time() * 1000),
    "appName": "PropNex",
}


def md5(data):
    h = hashlib.md5()
    h.update(bytes(data, encoding="utf-8"))
    return h.hexdigest()


def gen_signature(data):
    msg, sugar = "", "c1d65f3667324592a071ebec5038f38c"
    for i in sorted(data):
        if i not in ("file", "appVer", "mobileMode", "appSource", "token"):
            msg += str(data[i])
    return md5(msg + sugar)


def request(method, url, *, data=None, params=None):
    while True:
        try:
            if data:  
                data = urlencode(data)  
            return sess.request(method, url, data=data, params=params).json()
        except Exception as e:
            print(f"Request failed: {e}")
            time.sleep(0.01)

def login():

    email = os.getenv("email")
    password = os.getenv("password")
    

    body = {**params, "email": email, "password": md5(password)}
    body["signature"] = gen_signature(body)


    url = "https://api.singmap.com/app-service/agent/login"
    json = request("POST", url, data=body)
    # print(json)

    if json["success"]:
        print(Fore.GREEN + "Login successful" + Fore.RESET)
        params.update(
            {
                "token": json["datas"]["token"],
                "agentId": json["datas"]["agentId"],
                "brokeId": json["datas"]["brokeId"],
            }
        )
    else:
        print(Fore.RED + "Incorrect email or password. Please try again" + Fore.RESET)
        sys.exit()


def queryProjectList():
    body = {
        **params,
        "timestamp": int(time.time() * 1000),
        "pageSize": 1000,
        "pageNo": 1,
        "country": "Singapore",
    }
    body.update({"signature": gen_signature(body)})

    url = "https://api.singmap.com/app-service/project/queryProjectList"
    projects = request("POST", url, data=body)["datas"]["lists"]

    return {p["projectId"]: p["projectName"] for p in projects}


def queryFloorPlanList(projectId, projectName):
    query = {
        **params,
        "timestamp": int(time.time() * 1000),
        "projectId": projectId,
        "pageNo": "1",
        "pageSize": "200",
    }
    query.update({"signature": gen_signature(query)})

    url = "https://api.singmap.com/app-service/floor/queryFloorPlanList"
    planTypes = request("GET", url, params=query)["datas"]["lists"]

    if isinstance(planTypes, list):
        secondSheet(planTypes, projectId, projectName)


cost = re.compile(r"^\$((\d{0,3},?)+)$")
area = re.compile(r"^\d+/\d+$")


def queryUnitInfo(projectid, unitid):
    query = {
        **params,
        "timestamp": int(time.time() * 1000),
        "projectId": projectid,
        "unitId": unitid,
    }
    query.update({"signature": gen_signature(query)})

    url = "https://api.singmap.com/trade-service/unit/getUnitInfo"
    unit = request("GET", url, params=query)["datas"]

    info = {
        "unit": unit["unitName"],
        "price1": int(unit["price1"]),
        "price2": int(unit["price2"]),
        "area": int(unit["area"]),
        "price": int(unit["price"].replace(",", "") or 0),
        "psf": int(unit["avg"].replace(",", "") or 0),
    }
    for data in unit["lists"]:
        if isinstance(data["value"], str):
            if match := cost.match(data["value"]):
                info[data["key"]] = int(match[1].replace(",", ""))
                psqft = round((info["price2"] or info["price1"]) / info["area"])
                if psqft == info[data["key"]]:
                    info["psf"] = info[data["key"]]
            elif match := area.match(data["value"]):
                info[data["key"]] = int(match[0].split("/")[0])
            continue
        if isinstance(data["value"], (int, float)):
            info[data["key"]] = int(data["value"])

    min_price = min(info["price2"], info["price1"]) if info["price2"] and info["price1"] else info["price2"] or info["price1"]

    ws5.append(
        [
            unit["projectName"],
            unit["type"],
            unit["floorPlan"],
            unit["unitName"],
            min_price,
            info["psf"],
            info["area"],
        ]
    )
    return info


def queryUnitList(plan, projectId):
    body = {
        **params,
        "timestamp": int(time.time() * 1000),
        "projectId": projectId,
        "pageNo": "1",
        "pageSize": "1000",
        "orderType": "floorPlan",
        "floorPlan": plan,
    }
    body.update({"signature": gen_signature(body)})

    url = "https://api.singmap.com/app-service/unit/queryUnitList"
    units = request("POST", url, data=body)["datas"]["lists"]

    return [v["unitId"] for unit in units for v in unit["value"]]


def secondSheet(planTypes, projectId, projectName):
    ptu = pau = 0
    for planType in planTypes:
        tu = au = 0
        for plan in planType["list"]:
            au += plan["available"] or 0
            tu += plan["total"] or 0
        ptu += tu
        pau += au
        print(projectName, plan["floorPlanType"], tu, au)
        ws1.append([projectName, "", "", plan["floorPlanType"], tu, au])
    ws2.append([projectName, ptu, pau])

    thirdSheet(planTypes, projectId, projectName)


def thirdSheet(planTypes, projectId, projectName):
    for planType in planTypes:
        for plan in planType["list"]:
            ws3.append(
                [
                    projectName,
                    "",
                    "",
                    plan["floorPlanType"],
                    "",
                    "",
                    plan["floorPlanName"],
                    plan["total"] or 0,
                    plan["available"] or 0,
                ]
            )
            if plan["available"]:
                forthSheet(plan, projectId, projectName)


def forthSheet(plan, projectId, projectName):
    units = queryUnitList(plan["floorPlanName"], projectId)

    prices = set()
    psqfts = set()
    with concurrent.futures.ThreadPoolExecutor(max_workers=20) as ex:
        futures = {ex.submit(queryUnitInfo, projectId, unit) for unit in units}
        for future in futures:
            info = future.result()
            values = info.values()
            
            # Taking the min value out of two values of each unit.
            min_price = min(info["price2"], info["price1"]) if info["price2"] and info["price1"] else info["price2"] or info["price1"]

            if min_price != 0:
                prices.add(min_price)

            if (
                info["psf"] != 0
                and len(list(filter(lambda v: info["psf"] == v, values))) > 1
            ):
                psqfts.add(info["psf"])

    pricesl = list(prices)
    psqftsl = list(psqfts)

    if pricesl:
        min_price = min(pricesl)
        max_price = max(pricesl)
    else:
        min_price = max_price = 0

    if psqftsl:
        min_psqft = min(psqftsl)
        max_psqft = max(psqftsl)
    else:
        min_psqft = max_psqft = 0

    ws4.append([
        projectName,
        plan["floorPlanType"],
        plan["floorPlanName"],
        min_price,
        max_price,
        min_psqft,
        max_psqft
    ])
    

if __name__ == "__main__":
    login()

    for id, name in queryProjectList().items():
        queryFloorPlanList(id, name)

    # queryFloorPlanList(
    #     "c276c75b06c840bfa650f71b34ad67a9", "The LakeGarden Residences 嘉湖庭"
    # )

    wb1.save("Unit Name Available.xlsx")
    wb2.save("Project Available.xlsx")
    wb3.save("Unit Type Available.xlsx")
    wb4.save("Floor Plan Pricing Update.xlsx")
    wb5.save("Project Price List Update.xlsx")
